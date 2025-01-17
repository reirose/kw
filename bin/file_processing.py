import subprocess

import chardet
from langdetect import detect
from PIL import Image
import pytesseract
import openpyxl
import PyPDF2
# import os
from docx import Document
from bs4 import BeautifulSoup

# Путь к локальному файлу Tesseract
pytesseract.pytesseract.tesseract_cmd = r'.\Tesseract-OCR\tesseract.exe'


def detect_language(text):
    print(text)
    try:
        lang = detect(text)
        return lang
    except Exception as e:
        print("Error while detecting language:", e)
        return "unknown"


def detect_text_language(image_path):
    try:
        img = Image.open(image_path)
        text = pytesseract.image_to_string(img)
        return detect_language(text), text
    except Exception as e:
        print("Error while detecting text language:", e)
        return "unknown"


def extract_text_from_docx(file_path):
    try:
        document = Document(file_path)

        # Извлекаем текст из каждого параграфа
        text = []
        for paragraph in document.paragraphs:
            text.append(paragraph.text)

        # Объединяем текст в одну строку
        return "\n".join(text)
    except Exception as e:
        print("Error while extracting text from docx:", e)
        return ""


def extract_text_from_excel(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        text = ""
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        text += str(cell) + " "
                text += "\n"
        return text
    except Exception as e:
        print("Error while extracting text from Excel:", e)
        return ""


def extract_text_from_pdf(file_path):
    try:
        with open(file_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            text = ""
            for page_num in range(len(reader.pages)):
                page = reader.pages[page_num]
                text += page.extract_text()
            return text
    except Exception as e:
        print("Error while extracting text from PDF:", e)
        return ""


def extract_text_from_csv(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            text = file.read()
        return text
    except Exception as e:
        print("Error while extracting text from CSV:", e)
        return ""


def extract_text_from_html(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            soup = BeautifulSoup(file, 'html.parser')
            text = soup.get_text()
        return text
    except Exception as e:
        print("Error while extracting text from HTML:", e)
        return ""


def extract_text_from_txt(file_path):
    try:
        with open(file_path, 'rb') as file:
            raw_data = file.read()
            detected = chardet.detect(raw_data)
            encoding = detected['encoding']
        with open(file_path, 'r', encoding=encoding) as file:
            text = file.read()
        return text
    except Exception as e:
        print("Error while extracting text from txt:", e)
        return ""


def extract_text_from_doc(file_path):
    try:
        result = subprocess.run(['antiword', file_path], capture_output=True, text=True)
        if result.returncode == 0:
            return result.stdout

        result = subprocess.run(['catdoc', file_path], capture_output=True, text=True)
        if result.returncode == 0:
            return result.stdout

        raise RuntimeError("Both antiword and catdoc failed")
    except FileNotFoundError:
        raise RuntimeError("Please install antiword or catdoc: \n"
                           "For Ubuntu/Debian: sudo apt-get install antiword catdoc\n"
                           "For CentOS/RHEL: sudo yum install antiword catdoc\n"
                           "For Fedora: sudo dnf install antiword catdoc")



file_type_func = {
    "doc": extract_text_from_doc,
    "docx": extract_text_from_docx,
    "pdf": extract_text_from_pdf,
    "csv": extract_text_from_csv,
    "txt": extract_text_from_txt,
    "xlsx": extract_text_from_excel,
    "xls": extract_text_from_excel
}


def extract_text_and_language(file_path):
    file_type = file_path.split(".")[-1]
    if file_type not in file_type_func:
        try:
            text, lang = detect_text_language(file_path)
            return text, lang
        except Exception as e:
            raise TypeError("Неверный формат файла!")

    print(file_path)
    text = file_type_func[file_type](file_path)
    lang = detect_language(text)
    return text, lang
